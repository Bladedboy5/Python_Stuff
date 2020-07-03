import os
from openpyxl import Workbook, load_workbook
from statistics import mean

path  = 'C:\code\All_RFP\Test.xlsx'

if os.path.isfile(path):
    wb = load_workbook(path)
else:
    wb = Workbook()

ws = wb.active
row_max = ws.max_row
col_max = ws.max_column


class RFP:


    def DTT(self):
        b = []
        for j in range(1, col_max + 1):
            a = []
            for i in range(2, row_max + 1):
                cell = ws.cell(row = i, column = j)
                if cell.value == None:
                    cell.value = 0
                a.append(cell.value)
            b.append(mean(a))
            for i in b:
                ws.cell(row = row_max + 2, column = j).value = '%2f' % i

        ws.cell(row = row_max + 2, column = col_max + 2).value = '%2f' % min(b)
        ws.cell(row = row_max + 2, column = col_max + 3).value = '%2f' % max(b)

        wb.save(filename = 'C:\code\All_RFP\Test_Result.xlsx')

class RFP_Test:

    def fix(self):
        for j in range(1, col_max + 1):
            for i in range(2, row_max + 1):
                cell = ws.cell(row = i, column = j)
                if cell.value == str:
                    cell.value = 0
        wb.save(filename = 'C:\code\All_RFP\Test_Result.xlsx')

    def AVG(self):
        b = []
        for j in range(1, col_max + 1):
            a = []
            for i in range(2, row_max + 1):
                cell = ws.cell(row = i, column = j)
                a.append(cell.value)
            b.append(mean(a))
            for i in b:
                ws.cell(row = row_max + 2, column = j).value = int(i)
        wb.save(filename = 'C:\code\All_RFP\Test.xlsx')

    def max_min(self):
        b = []
        for j in range(1, col_max + 1):
            a = []
            for i in range(2, row_max + 1):
                cell = ws.cell(row = i, column = j)
                a.append(cell.value)
            b.append(mean(a))
        ws.cell(row = row_max + 2, column = col_max + 2).value = int(min(b))
        ws.cell(row = row_max + 2, column = col_max + 3).value = int(max(b))
        wb.save(filename = 'C:\code\All_RFP\Test.xlsx')
        